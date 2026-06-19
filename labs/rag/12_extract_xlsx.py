"""
12_extract_xlsx.py

XLSX 문서에서 Sheet와 Row 단위로 데이터를 추출한다.

입력:
- enterprise_docs/xlsx/*.xlsx

출력:
- extracted_text/xlsx_extracted.jsonl

실행:
python 12_extract_xlsx.py
"""

from __future__ import annotations

from openpyxl import load_workbook

from step2_5_config import ENTERPRISE_DOCS_DIR, EXTRACTED_TEXT_DIR, ensure_directories
from step2_5_utils import clean_text, find_files, stable_id, write_jsonl, print_record_summary


def normalize_cell(value) -> str:
    """Excel 셀 값을 문자열로 변환한다."""
    if value is None:
        return ""
    return clean_text(str(value))


def extract_xlsx_files() -> list[dict]:
    ensure_directories()

    xlsx_dir = ENTERPRISE_DOCS_DIR / "xlsx"
    xlsx_files = find_files(xlsx_dir, (".xlsx",))

    records: list[dict] = []

    for xlsx_path in xlsx_files:
        try:
            workbook = load_workbook(str(xlsx_path), data_only=True)
        except Exception as exc:
            print(f"[XLSX 읽기 실패] {xlsx_path.name}: {exc}")
            continue

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                continue

            # 첫 번째 행을 헤더로 가정한다.
            headers = [normalize_cell(value) for value in rows[0]]

            for row_no, row in enumerate(rows[1:], start=2):
                values: list[str] = []

                for col_index, value in enumerate(row):
                    cell_value = normalize_cell(value)

                    if not cell_value:
                        continue

                    header = headers[col_index] if col_index < len(headers) and headers[col_index] else f"컬럼{col_index + 1}"
                    values.append(f"{header}: {cell_value}")

                if not values:
                    continue

                row_text = clean_text(", ".join(values))

                records.append(
                    {
                        "id": stable_id(xlsx_path.name, "xlsx", sheet.title, row_no),
                        "text": row_text,
                        "metadata": {
                            "file_name": xlsx_path.name,
                            "document_type": "xlsx",
                            "sheet_name": sheet.title,
                            "row_no": row_no,
                            "source_path": str(xlsx_path),
                            "extractor": "openpyxl",
                        },
                    }
                )

    return records


def main() -> None:
    output_path = EXTRACTED_TEXT_DIR / "xlsx_extracted.jsonl"
    records = extract_xlsx_files()
    count = write_jsonl(output_path, records)
    print_record_summary("XLSX 텍스트 추출 완료", count, output_path)


if __name__ == "__main__":
    main()
